import warnings
warnings.filterwarnings('ignore', category=FutureWarning)

import subprocess
import time
import win32com.client
import os
import datetime
import calendar
import openpyxl
import gspread
import sys
import requests
import win32clipboard
from google.auth.transport.requests import AuthorizedSession
from google.oauth2 import service_account

OUTPUT_DIR_IH08 = r"O:\My Drive\071 - SAP 数据\A_Equipment"
OUTPUT_DIR_IP18 = r"O:\My Drive\071 - SAP 数据\A_Equipments_with_Maintenance_Plan"
GOOGLE_SHEET_ID = '1dWWNiIlRqnXDeDiY4MS_cPSfseDb0g1NHK1dC1HPSnY'
WORKSHEET_NAME = 'MasterData'

def get_resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

SERVICE_ACCOUNT_FILE = get_resource_path('pyreadsp-b5b9c1909de6.json')

def sap_auto_logo():
    subprocess.check_call(['C:\\Program Files (x86)\\SAP\\FrontEnd\\SAPgui\\sapshcut.exe', '-system=CAP', '-client=321',
                           '-user=USERNAME', '-pw=PASSWORD', '-language=ZH'])
    time.sleep(15)
    print("sap open successfully")

def close_SAP():
    try:
        os.system('taskkill /im saplogon.exe /t /f')
        print("SAP GUI 进程已成功关闭。")
    except Exception as e:
        print(f"关闭 SAP 进程失败: {e}")

def run_ih08(session, year_month, first_day_str, last_day_str):
    output_filename = f"{year_month}_A_Equipments.xlsx"
    output_file_path = os.path.join(OUTPUT_DIR_IH08, output_filename)

    if os.path.exists(output_file_path):
        try:
            os.remove(output_file_path)
            print(f"  已删除旧文件: {output_file_path}")
        except Exception as e:
            print(f"  警告: 无法删除旧文件 {output_file_path}: {e}")

    print("  执行 IH08...")
    session.findById("wnd[0]").maximize()
    session.findById("wnd[0]/tbar[0]/okcd").text = "IH08"
    session.findById("wnd[0]").sendVKey(0)
    time.sleep(1)

    session.findById("wnd[0]/usr/ctxtDATUV").text = first_day_str
    session.findById("wnd[0]/usr/ctxtDATUB").text = last_day_str
    session.findById("wnd[0]/usr/ctxtDATUB").setFocus()
    session.findById("wnd[0]/usr/ctxtDATUB").caretPosition = 10

    session.findById("wnd[0]/usr/btn%_STAE1_%_APP_%-VALU_PUSH").press()
    time.sleep(1)

    session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,0]").text = "ASEQ"
    session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,1]").text = "DLFL"
    session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,2]").text = "DLT"
    session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,3]").text = "INAC"
    session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,4]").text = "标记"
    session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,4]").setFocus()
    session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,4]").caretPosition = 2
    session.findById("wnd[1]/tbar[0]/btn[8]").press()
    time.sleep(1)

    session.findById("wnd[0]/usr/ctxtSWERK-LOW").text = "CN15"
    session.findById("wnd[0]/usr/ctxtABCKZ-LOW").text = "A"
    session.findById("wnd[0]/usr/ctxtVARIANT").text = "/GUAN_EQUIP"
    session.findById("wnd[0]/usr/ctxtVARIANT").setFocus()
    session.findById("wnd[0]/usr/ctxtVARIANT").caretPosition = 11

    session.findById("wnd[0]/tbar[1]/btn[8]").press()
    time.sleep(3)

    session.findById("wnd[0]/tbar[1]/btn[16]").press()
    time.sleep(1)

    session.findById("wnd[1]/tbar[0]/btn[0]").press()
    session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[0,0]").select()
    session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[0,0]").setFocus()
    session.findById("wnd[1]/tbar[0]/btn[0]").press()
    session.findById("wnd[1]/tbar[0]/btn[0]").press()

    print("  等待 Excel 打开...")
    time.sleep(5)

    try:
        excel = win32com.client.Dispatch("Excel.Application")
        if excel.Workbooks.Count > 0:
            wb = excel.Workbooks(1)
            print(f"  找到打开的工作簿: {wb.Name}")
            if not os.path.exists(OUTPUT_DIR_IH08):
                os.makedirs(OUTPUT_DIR_IH08)
            wb.SaveAs(output_file_path)
            wb.Close(SaveChanges=False)
            excel.Quit()
            print(f"  ✅ IH08 导出完成: {output_filename}")
        else:
            print("  ⚠️ 警告: 没有找到打开的 Excel 工作簿")
            return None
    except Exception as e:
        print(f"  ❌ 处理 Excel 时发生错误: {e}")
        return None

    session.findById("wnd[0]").close()
    time.sleep(1)
    return output_file_path


def run_ip18(session, year_month, ih08_file):
    output_filename = f"{year_month}_A_Equipment_with_Maintenance_Plan.xlsx"
    output_file_path = os.path.join(OUTPUT_DIR_IP18, output_filename)

    if os.path.exists(output_file_path):
        try:
            os.remove(output_file_path)
            print(f"  已删除旧文件: {output_file_path}")
        except Exception as e:
            print(f"  警告: 无法删除旧文件 {output_file_path}: {e}")

    print("  从 IH08 文件读取设备编号...")
    try:
        wb_ih08 = openpyxl.load_workbook(ih08_file)
        ws_ih08 = wb_ih08.active
        equipment_col = None
        for col in range(1, ws_ih08.max_column + 1):
            if ws_ih08.cell(1, col).value == "设备":
                equipment_col = col
                break
        if not equipment_col:
            print("  ❌ IH08 文件中未找到'设备'列")
            wb_ih08.close()
            return None
        equipment_list = []
        for row in range(2, ws_ih08.max_row + 1):
            equipment = ws_ih08.cell(row, equipment_col).value
            if equipment:
                equipment_list.append(str(equipment).strip())
        wb_ih08.close()
        print(f"  从 IH08 读取到 {len(equipment_list)} 个设备编号")
    except Exception as e:
        print(f"  ❌ 读取 IH08 文件失败: {e}")
        return None

    print("  执行 IP18...")
    session.findById("wnd[0]/tbar[0]/okcd").text = "/NIP18"
    session.findById("wnd[0]").sendVKey(0)
    time.sleep(1)

    session.findById("wnd[0]/usr/btn%_EQUNR_%_APP_%-VALU_PUSH").press()
    time.sleep(1)

    clipboard_text = "\r\n".join(equipment_list)
    win32clipboard.OpenClipboard()
    win32clipboard.EmptyClipboard()
    win32clipboard.SetClipboardText(clipboard_text, win32clipboard.CF_UNICODETEXT)
    win32clipboard.CloseClipboard()

    session.findById("wnd[1]/tbar[0]/btn[24]").press()
    time.sleep(2)

    session.findById("wnd[1]/tbar[0]/btn[8]").press()
    time.sleep(1)

    session.findById("wnd[0]/usr/chkSPERRE").selected = True
    session.findById("wnd[0]/usr/ctxtIWERK-LOW").text = "CN15"
    session.findById("wnd[0]/usr/ctxtSWERK-LOW").text = "CN15"
    session.findById("wnd[0]/usr/ctxtSWERK-LOW").setFocus()
    session.findById("wnd[0]/usr/ctxtSWERK-LOW").caretPosition = 4

    session.findById("wnd[0]/tbar[1]/btn[8]").press()
    time.sleep(3)

    session.findById("wnd[0]/mbar/menu[0]/menu[5]").select()
    time.sleep(1)

    session.findById("wnd[1]/tbar[0]/btn[0]").press()
    session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[0,0]").select()
    session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[0,0]").setFocus()
    session.findById("wnd[1]/tbar[0]/btn[0]").press()
    session.findById("wnd[1]/tbar[0]/btn[0]").press()

    print("  等待 Excel 打开...")
    time.sleep(5)

    try:
        excel = win32com.client.Dispatch("Excel.Application")
        if excel.Workbooks.Count > 0:
            wb = excel.Workbooks(1)
            print(f"  找到打开的工作簿: {wb.Name}")
            if not os.path.exists(OUTPUT_DIR_IP18):
                os.makedirs(OUTPUT_DIR_IP18)
            wb.SaveAs(output_file_path)
            wb.Close(SaveChanges=False)
            excel.Quit()
            print(f"  ✅ IP18 导出完成: {output_filename}")
        else:
            print("  ⚠️ 警告: 没有找到打开的 Excel 工作簿")
            return None
    except Exception as e:
        print(f"  ❌ 处理 Excel 时发生错误: {e}")
        return None

    return output_file_path


def process_data(ih08_file, ip18_file, year_month):
    wb_ih08 = openpyxl.load_workbook(ih08_file)
    ws_ih08 = wb_ih08.active

    equipment_col_ih08 = None
    desc_col_ih08 = None
    for col in range(1, ws_ih08.max_column + 1):
        header = ws_ih08.cell(1, col).value
        if header == "设备":
            equipment_col_ih08 = col
        elif header and "描述" in str(header):
            desc_col_ih08 = col

    equipments_ih08 = set()
    equipment_desc = {}
    for row in range(2, ws_ih08.max_row + 1):
        equipment = ws_ih08.cell(row, equipment_col_ih08).value
        if equipment:
            eq_str = str(equipment).strip()
            equipments_ih08.add(eq_str)
            if desc_col_ih08:
                desc = ws_ih08.cell(row, desc_col_ih08).value
                equipment_desc[eq_str] = str(desc).strip() if desc else ""
    wb_ih08.close()
    total_equipments = len(equipments_ih08)

    wb_ip18 = openpyxl.load_workbook(ip18_file)
    ws_ip18 = wb_ip18.active
    equipment_col_ip18 = None
    for col in range(1, ws_ip18.max_column + 1):
        if ws_ip18.cell(1, col).value == "设备":
            equipment_col_ip18 = col
            break

    equipments_ip18 = set()
    for row in range(2, ws_ip18.max_row + 1):
        equipment = ws_ip18.cell(row, equipment_col_ip18).value
        if equipment:
            equipments_ip18.add(str(equipment).strip())
    wb_ip18.close()
    equipments_with_plan = len(equipments_ip18)

    equipments_without_plan = sorted(equipments_ih08 - equipments_ip18)
    equipments_without_plan_data = [
        [eq, equipment_desc.get(eq, ""), year_month]
        for eq in equipments_without_plan
    ]
    percentage = round((equipments_with_plan / total_equipments), 4) if total_equipments > 0 else 0.0

    return {
        'month': year_month,
        'equipments_with_plan': equipments_with_plan,
        'total_equipments': total_equipments,
        'percentage': percentage,
        'equipments_without_plan': equipments_without_plan_data
    }


def upload_to_master(data, sheet_id, worksheet_name, auth_file):
    session = requests.Session()
    session.verify = False
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    credentials = service_account.Credentials.from_service_account_file(
        auth_file, scopes=['https://www.googleapis.com/auth/spreadsheets'])
    authed_session = AuthorizedSession(credentials)
    authed_session.verify = False
    gc = gspread.Client(auth=credentials, session=authed_session)
    sh = gc.open_by_key(sheet_id)
    worksheet = sh.worksheet(worksheet_name)

    all_values = worksheet.get_all_values()
    target_row = None
    for idx, row in enumerate(all_values):
        if len(row) > 0 and row[0] == data['month']:
            target_row = idx + 1
            break

    update_data = [data['equipments_with_plan'], data['total_equipments'], data['percentage']]
    if target_row:
        worksheet.update(f'H{target_row}:J{target_row}', [update_data])
        print(f"  ✅ 已更新月份 {data['month']} 的数据（第 {target_row} 行，H-J 列）")
    else:
        new_row = [data['month'], '', '', '', '', '', ''] + update_data
        worksheet.append_row(new_row, value_input_option='USER_ENTERED')
        print(f"  ✅ 已添加月份 {data['month']} 的新行（H-J 列）")
    return True


def upload_no_plan(equipments, sheet_id, auth_file):
    session = requests.Session()
    session.verify = False
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    credentials = service_account.Credentials.from_service_account_file(
        auth_file, scopes=['https://www.googleapis.com/auth/spreadsheets'])
    authed_session = AuthorizedSession(credentials)
    authed_session.verify = False
    gc = gspread.Client(auth=credentials, session=authed_session)
    sh = gc.open_by_key(sheet_id)

    sheet_name = '无保养计划A类设备'
    try:
        worksheet = sh.worksheet(sheet_name)
    except gspread.exceptions.WorksheetNotFound:
        worksheet = sh.add_worksheet(title=sheet_name, rows="1000", cols="3")
        worksheet.update('A1:C1', [['设备编号', '描述', '月份']])

    if equipments:
        worksheet.append_rows(equipments, value_input_option='USER_ENTERED')
    print(f"  ✅ 已上传 {len(equipments)} 个无保养计划A类设备")
    return True


if __name__ == "__main__":
    months = ["202601", "202602", "202603", "202604", "202605", "202606"]

    print("正在清理可能存在的 SAP 进程...")
    close_SAP()
    time.sleep(2)

    print("正在启动 SAP GUI...")
    sap_auto_logo()
    time.sleep(5)

    try:
        SapGuiAuto = win32com.client.GetObject("SAPGUI")
        application = SapGuiAuto.GetScriptingEngine
        connection = application.Children(0)
        session = connection.Children(0)
        print("成功连接到 SAP 会话。")
    except Exception as e:
        print(f"错误: 无法连接到 SAP: {e}")
        close_SAP()
        input()
        sys.exit(1)

    for ym in months:
        print(f"\n{'='*50}")
        print(f"处理 {ym}")
        print(f"{'='*50}")

        year = int(ym[:4])
        month = int(ym[4:6])
        first_day = datetime.date(year, month, 1)
        last_day = datetime.date(year, month, calendar.monthrange(year, month)[1])
        first_day_str = first_day.strftime("%m/%d/%Y")
        last_day_str = last_day.strftime("%m/%d/%Y")

        ih08_file = run_ih08(session, ym, first_day_str, last_day_str)
        if not ih08_file:
            print(f"  ❌ {ym} IH08 失败，跳过")
            continue

        ip18_file = run_ip18(session, ym, ih08_file)
        if not ip18_file:
            print(f"  ❌ {ym} IP18 失败，跳过")
            continue

        data = process_data(ih08_file, ip18_file, ym)
        print(f"  有保养计划: {data['equipments_with_plan']}/{data['total_equipments']}, 百分比: {data['percentage']}")

        upload_to_master(data, GOOGLE_SHEET_ID, WORKSHEET_NAME, SERVICE_ACCOUNT_FILE)

        if data['equipments_without_plan']:
            print(f"  无保养计划设备: {len(data['equipments_without_plan'])} 个")
            upload_no_plan(data['equipments_without_plan'], GOOGLE_SHEET_ID, SERVICE_ACCOUNT_FILE)

    print("\nSAP 操作完成，正在关闭 SAP...")
    time.sleep(2)
    close_SAP()

    print("\n" + "="*50)
    print("✅ 全部完成！")
    print("="*50)
    input()
