import subprocess
import time
import win32com.client
import os
import datetime
import calendar
import openpyxl
import sys
import warnings
import win32clipboard

warnings.filterwarnings("ignore", category=FutureWarning)
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

import gspread
from google.auth.transport.requests import AuthorizedSession
from google.oauth2 import service_account
import requests

OUTPUT_DIR_IH08 = r"O:\My Drive\071 - SAP 数据\A_Equipment"
OUTPUT_DIR_IP18 = r"O:\My Drive\071 - SAP 数据\A_Equipments_with_Maintenance_Plan"

GOOGLE_SHEET_ID = '1dWWNiIlRqnXDeDiY4MS_cPSfseDb0g1NHK1dC1HPSnY'
WORKSHEET_NAME = 'MasterData'

def get_resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

SERVICE_ACCOUNT_FILE = get_resource_path('pyreadsp-b5b9c1909de6.json')

def sap_auto_logo():
    subprocess.check_call(['C:\\Program Files (x86)\\SAP\\FrontEnd\\SAPgui\\sapshcut.exe', '-system=CAP', '-client=321',
                           '-user=USERNAME', '-pw=PASSWORD', '-language=ZH'])
    time.sleep(15)
    print("sap open successfully")


def close_SAP():
    try:
        os.system('taskkill /im saplogon.exe /t /f')
        print("SAP GUI 进程已成功关闭。")
    except Exception as e:
        print(f"关闭 SAP 进程失败: {e}")


def get_a_equipments_ih08(session, year_month, first_day_str, last_day_str):
    output_filename = f"{year_month}_A_Equipments.xlsx"
    output_file_path = os.path.join(OUTPUT_DIR_IH08, output_filename)
    
    if os.path.exists(output_file_path):
        try:
            os.remove(output_file_path)
            print(f"已删除旧文件: {output_file_path}")
        except Exception as e:
            print(f"警告: 无法删除旧文件 {output_file_path}: {e}")
    
    print("正在执行 IH08 - 获取 A 级关键设备...")
    
    try:
        session.findById("wnd[0]").maximize()
        session.findById("wnd[0]/tbar[0]/okcd").text = "IH08"
        session.findById("wnd[0]").sendVKey(0)
        time.sleep(1)
        
        session.findById("wnd[0]/usr/ctxtDATUV").text = first_day_str
        session.findById("wnd[0]/usr/ctxtDATUB").text = last_day_str
        session.findById("wnd[0]/usr/ctxtDATUB").setFocus()
        session.findById("wnd[0]/usr/ctxtDATUB").caretPosition = 10
        
        session.findById("wnd[0]/usr/btn%_STAE1_%_APP_%-VALU_PUSH").press()
        time.sleep(1)
        
        session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,0]").text = "ASEQ"
        session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,1]").text = "DLFL"
        session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,2]").text = "DLT"
        session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,3]").text = "INAC"
        session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,4]").text = "标记"
        session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,4]").setFocus()
        session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,4]").caretPosition = 2
        session.findById("wnd[1]/tbar[0]/btn[8]").press()
        time.sleep(1)
        
        session.findById("wnd[0]/usr/ctxtSWERK-LOW").text = "CN15"
        session.findById("wnd[0]/usr/ctxtABCKZ-LOW").text = "A"
        session.findById("wnd[0]/usr/ctxtVARIANT").text = "/GUAN_EQUIP"
        session.findById("wnd[0]/usr/ctxtVARIANT").setFocus()
        session.findById("wnd[0]/usr/ctxtVARIANT").caretPosition = 11
        
        session.findById("wnd[0]/tbar[1]/btn[8]").press()
        time.sleep(3)
        
        session.findById("wnd[0]/tbar[1]/btn[16]").press()
        time.sleep(1)
        
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        
        session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[0,0]").select()
        session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[0,0]").setFocus()
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        
        print("等待 Excel 打开...")
        time.sleep(5)
        
        print("正在连接到 Excel 应用程序...")
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            
            if excel.Workbooks.Count > 0:
                wb = excel.Workbooks(1)
                print(f"找到打开的工作簿: {wb.Name}")
                
                if not os.path.exists(OUTPUT_DIR_IH08):
                    os.makedirs(OUTPUT_DIR_IH08)
                    print(f"创建目录: {OUTPUT_DIR_IH08}")
                
                print(f"正在另存为到: {output_file_path}")
                wb.SaveAs(output_file_path)
                
                wb.Close(SaveChanges=False)
                excel.Quit()
                
                print(f"✅ IH08 A 级设备数据导出完成")
                print(f"文件保存路径: {output_file_path}")
                
            else:
                print("⚠️ 警告: 没有找到打开的 Excel 工作簿")
                return None
        except Exception as e:
            print(f"❌ 处理 Excel 时发生错误: {e}")
            import traceback
            traceback.print_exc()
            return None
        
        session.findById("wnd[0]").close()
        time.sleep(1)
        
        return output_file_path
        
    except Exception as e:
        print(f"❌ 执行 IH08 时发生错误: {e}")
        import traceback
        traceback.print_exc()
        return None


def get_equipments_with_plan_ip18(session, year_month, ih08_file):
    output_filename = f"{year_month}_A_Equipment_with_Maintenance_Plan.xlsx"
    output_file_path = os.path.join(OUTPUT_DIR_IP18, output_filename)
    
    if os.path.exists(output_file_path):
        try:
            os.remove(output_file_path)
            print(f"已删除旧文件: {output_file_path}")
        except Exception as e:
            print(f"警告: 无法删除旧文件 {output_file_path}: {e}")
    
    print("\n正在从 IH08 文件读取设备编号...")
    try:
        wb_ih08 = openpyxl.load_workbook(ih08_file)
        ws_ih08 = wb_ih08.active
        
        equipment_col = None
        for col in range(1, ws_ih08.max_column + 1):
            if ws_ih08.cell(1, col).value == "设备":
                equipment_col = col
                break
        
        if not equipment_col:
            print("❌ IH08 文件中未找到'设备'列")
            wb_ih08.close()
            return None
        
        equipment_list = []
        for row in range(2, ws_ih08.max_row + 1):
            equipment = ws_ih08.cell(row, equipment_col).value
            if equipment:
                equipment_list.append(str(equipment).strip())
        
        wb_ih08.close()
        print(f"从 IH08 读取到 {len(equipment_list)} 个设备编号")
        
    except Exception as e:
        print(f"❌ 读取 IH08 文件失败: {e}")
        return None
    
    print("\n正在执行 IP18 - 获取有维护计划的设备...")
    
    try:
        session.findById("wnd[0]/tbar[0]/okcd").text = "/NIP18"
        session.findById("wnd[0]").sendVKey(0)
        time.sleep(1)
        
        session.findById("wnd[0]/usr/btn%_EQUNR_%_APP_%-VALU_PUSH").press()
        time.sleep(1)
        
        print(f"正在将 {len(equipment_list)} 个设备编号复制到剪贴板...")
        clipboard_text = "\r\n".join(equipment_list)
        win32clipboard.OpenClipboard()
        win32clipboard.EmptyClipboard()
        win32clipboard.SetClipboardText(clipboard_text, win32clipboard.CF_UNICODETEXT)
        win32clipboard.CloseClipboard()
        print(f"✅ 已将 {len(equipment_list)} 个设备编号复制到剪贴板")
        
        print("正在点击'自剪贴板上载'按钮...")
        session.findById("wnd[1]/tbar[0]/btn[24]").press()
        time.sleep(2)
        print(f"✅ 设备编号批量上载完成")
        
        session.findById("wnd[1]/tbar[0]/btn[8]").press()
        time.sleep(1)
        
        session.findById("wnd[0]/usr/chkSPERRE").selected = True
        session.findById("wnd[0]/usr/ctxtIWERK-LOW").text = "CN15"
        session.findById("wnd[0]/usr/ctxtSWERK-LOW").text = "CN15"
        session.findById("wnd[0]/usr/ctxtSWERK-LOW").setFocus()
        session.findById("wnd[0]/usr/ctxtSWERK-LOW").caretPosition = 4
        
        session.findById("wnd[0]/tbar[1]/btn[8]").press()
        time.sleep(3)
        
        session.findById("wnd[0]/mbar/menu[0]/menu[5]").select()
        time.sleep(1)
        
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        
        session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[0,0]").select()
        session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[0,0]").setFocus()
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        
        print("等待 Excel 打开...")
        time.sleep(5)
        
        print("正在连接到 Excel 应用程序...")
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            
            if excel.Workbooks.Count > 0:
                wb = excel.Workbooks(1)
                print(f"找到打开的工作簿: {wb.Name}")
                
                if not os.path.exists(OUTPUT_DIR_IP18):
                    os.makedirs(OUTPUT_DIR_IP18)
                    print(f"创建目录: {OUTPUT_DIR_IP18}")
                
                print(f"正在另存为到: {output_file_path}")
                wb.SaveAs(output_file_path)
                
                wb.Close(SaveChanges=False)
                excel.Quit()
                
                print(f"✅ IP18 有维护计划的设备数据导出完成")
                print(f"文件保存路径: {output_file_path}")
                
            else:
                print("⚠️ 警告: 没有找到打开的 Excel 工作簿")
                return None
        except Exception as e:
            print(f"❌ 处理 Excel 时发生错误: {e}")
            import traceback
            traceback.print_exc()
            return None
        
        return output_file_path
        
    except Exception as e:
        print(f"❌ 执行 IP18 时发生错误: {e}")
        import traceback
        traceback.print_exc()
        return None


def process_critical_equipment_data(ih08_file, ip18_file, year_month):
    print("\n" + "="*50)
    print("开始处理数据...")
    print("="*50)
    
    try:
        wb_ih08 = openpyxl.load_workbook(ih08_file)
        ws_ih08 = wb_ih08.active
        
        equipment_col_ih08 = None
        desc_col_ih08 = None
        for col in range(1, ws_ih08.max_column + 1):
            header = ws_ih08.cell(1, col).value
            if header == "设备":
                equipment_col_ih08 = col
            elif header and "描述" in str(header):
                desc_col_ih08 = col
        
        if not equipment_col_ih08:
            print("❌ IH08 文件中未找到'设备'列")
            wb_ih08.close()
            return None
        
        equipments_ih08 = set()
        equipment_desc = {}
        for row in range(2, ws_ih08.max_row + 1):
            equipment = ws_ih08.cell(row, equipment_col_ih08).value
            if equipment:
                eq_str = str(equipment).strip()
                equipments_ih08.add(eq_str)
                if desc_col_ih08:
                    desc = ws_ih08.cell(row, desc_col_ih08).value
                    equipment_desc[eq_str] = str(desc).strip() if desc else ""
        
        wb_ih08.close()
        total_equipments = len(equipments_ih08)
        print(f"IH08 - A 级关键设备总数（去重）: {total_equipments}")
        
        wb_ip18 = openpyxl.load_workbook(ip18_file)
        ws_ip18 = wb_ip18.active
        
        equipment_col_ip18 = None
        for col in range(1, ws_ip18.max_column + 1):
            if ws_ip18.cell(1, col).value == "设备":
                equipment_col_ip18 = col
                break
        
        if not equipment_col_ip18:
            print("❌ IP18 文件中未找到'设备'列")
            wb_ip18.close()
            return None
        
        equipments_ip18 = set()
        for row in range(2, ws_ip18.max_row + 1):
            equipment = ws_ip18.cell(row, equipment_col_ip18).value
            if equipment:
                equipments_ip18.add(str(equipment).strip())
        
        wb_ip18.close()
        equipments_with_plan = len(equipments_ip18)
        print(f"IP18 - 有维护计划的设备总数（去重）: {equipments_with_plan}")
        
        equipments_without_plan = sorted(equipments_ih08 - equipments_ip18)
        equipments_without_plan_data = [
            [eq, equipment_desc.get(eq, ""), year_month]
            for eq in equipments_without_plan
        ]
        percentage = round((equipments_with_plan / total_equipments), 4) if total_equipments > 0 else 0.0
        
        return {
            'month': year_month,
            'equipments_with_plan': equipments_with_plan,
            'total_equipments': total_equipments,
            'percentage': percentage,
            'equipments_without_plan': equipments_without_plan_data
        }
        
    except Exception as e:
        print(f"❌ 处理数据时发生错误: {e}")
        import traceback
        traceback.print_exc()
        return None


def upload_to_google_sheets(data, sheet_id, worksheet_name, auth_file):
    try:
        print("\n正在连接 Google Sheets...")
        
        session = requests.Session()
        session.verify = False
        
        credentials = service_account.Credentials.from_service_account_file(
            auth_file,
            scopes=['https://www.googleapis.com/auth/spreadsheets']
        )
        
        authed_session = AuthorizedSession(credentials)
        authed_session.verify = False
        
        gc = gspread.Client(auth=credentials, session=authed_session)
        sh = gc.open_by_key(sheet_id)
        worksheet = sh.worksheet(worksheet_name)
        
        all_values = worksheet.get_all_values()
        month_col_index = 0
        
        target_row = None
        for idx, row in enumerate(all_values):
            if len(row) > 0 and row[month_col_index] == data['month']:
                target_row = idx + 1
                break
        
        update_data = [data['equipments_with_plan'], data['total_equipments'], data['percentage']]
        if target_row:
            worksheet.update(f'H{target_row}:J{target_row}', [update_data])
            print(f"✅ 已更新月份 {data['month']} 的数据（第 {target_row} 行，H-J 列）")
        else:
            new_row = [data['month'], '', '', '', '', '', ''] + update_data
            worksheet.append_row(new_row, value_input_option='USER_ENTERED')
            print(f"✅ 已添加月份 {data['month']} 的新行（H-J 列）")
        
        return True
        
    except gspread.exceptions.WorksheetNotFound:
        print(f"❌ Google Sheets 错误: 找不到工作表 '{worksheet_name}'。")
        return False
    except gspread.exceptions.SpreadsheetNotFound:
        print(f"❌ Google Sheets 错误: 找不到工作簿。请确认已将服务账户邮箱添加为该表格的 '编辑者'。")
        return False
    except Exception as e:
        print(f"❌ Google Sheets 上传失败。错误: {e}")
        import traceback
        traceback.print_exc()
        return False


def upload_no_plan_equipments(equipments, sheet_id, auth_file):
    try:
        print("\n正在上传无保养计划A类设备清单...")
        session = requests.Session()
        session.verify = False
        credentials = service_account.Credentials.from_service_account_file(
            auth_file,
            scopes=['https://www.googleapis.com/auth/spreadsheets']
        )
        authed_session = AuthorizedSession(credentials)
        authed_session.verify = False
        gc = gspread.Client(auth=credentials, session=authed_session)
        sh = gc.open_by_key(sheet_id)

        sheet_name = '无保养计划A类设备'
        try:
            worksheet = sh.worksheet(sheet_name)
        except gspread.exceptions.WorksheetNotFound:
            worksheet = sh.add_worksheet(title=sheet_name, rows="1000", cols="3")
            worksheet.update('A1:C1', [['设备编号', '描述', '月份']])

        if equipments:
            all_rows = worksheet.get_all_values()
            year_month = equipments[0][2]
            keep_rows = [all_rows[0]] if all_rows else []
            for row in all_rows[1:]:
                if len(row) < 3 or row[2] != year_month:
                    keep_rows.append(row)
            if len(keep_rows) < len(all_rows):
                worksheet.clear()
                worksheet.update(keep_rows, 'A1', value_input_option='USER_ENTERED')
                print(f"  已清除当月旧数据 {len(all_rows) - len(keep_rows)} 行")
            worksheet.append_rows(equipments, value_input_option='USER_ENTERED')

        print(f"✅ 已上传 {len(equipments)} 个无保养计划A类设备到 '{sheet_name}'")
        return True
    except Exception as e:
        print(f"❌ 上传无保养计划设备清单失败: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    try:
        print("正在清理可能存在的 SAP 进程...")
        close_SAP()
        time.sleep(2)
        
        print("正在启动 SAP GUI...")
        sap_auto_logo()
        
        print("等待 SAP GUI 完全加载...")
        time.sleep(5)
        
        try:
            SapGuiAuto = win32com.client.GetObject("SAPGUI")
            application = SapGuiAuto.GetScriptingEngine
            connection = application.Children(0)
            session = connection.Children(0)
        except Exception as e:
            print(f"错误: 无法连接到 SAP GUI Scripting Engine 或找不到活动会话。请确保 SAP GUI 已登录。{e}")
            close_SAP()
            input("\n按任意键退出...")
            exit(1)
        
        print("成功连接到 SAP 会话。")
        
        today = datetime.date.today()
        first_day = today.replace(day=1)
        last_day = today.replace(day=calendar.monthrange(today.year, today.month)[1])
        
        first_day_str = first_day.strftime("%m/%d/%Y")
        last_day_str = last_day.strftime("%m/%d/%Y")
        year_month = today.strftime("%Y%m")
        
        print(f"\n日期范围: {first_day_str} - {last_day_str}")
        print(f"月份: {year_month}")
        
        ih08_file = get_a_equipments_ih08(session, year_month, first_day_str, last_day_str)
        
        if not ih08_file:
            print("❌ IH08 执行失败，终止程序")
            close_SAP()
            input("\n按任意键退出...")
            exit(1)
        
        ip18_file = get_equipments_with_plan_ip18(session, year_month, ih08_file)
        
        if not ip18_file:
            print("❌ IP18 执行失败，终止程序")
            close_SAP()
            input("\n按任意键退出...")
            exit(1)
        
        data = process_critical_equipment_data(ih08_file, ip18_file, year_month)
        
        if data:
            print(f"\n统计结果：")
            print(f"  - 月份: {data['month']}")
            print(f"  - 有维护计划的设备数: {data['equipments_with_plan']}")
            print(f"  - A 级关键设备总数: {data['total_equipments']}")
            print(f"  - 百分比: {data['percentage']}")
            
            upload_to_google_sheets(
                data=data,
                sheet_id=GOOGLE_SHEET_ID,
                worksheet_name=WORKSHEET_NAME,
                auth_file=SERVICE_ACCOUNT_FILE
            )

            if data['equipments_without_plan']:
                print(f"\n无保养计划的A类设备数量: {len(data['equipments_without_plan'])}")
                upload_no_plan_equipments(
                    equipments=data['equipments_without_plan'],
                    sheet_id=GOOGLE_SHEET_ID,
                    auth_file=SERVICE_ACCOUNT_FILE
                )
            else:
                print("\n✅ 所有A类设备均有保养计划")
        else:
            print("❌ 数据处理失败")
        
        print("\nSAP 操作完成，正在关闭 SAP...")
        close_SAP()
        
        print("\n" + "="*50)
        print("✅ 所有操作已完成！")
        print("="*50)
        
    except Exception as e:
        print(f"❌ 程序执行过程中发生错误: {e}")
        import traceback
        traceback.print_exc()
        close_SAP()
    finally:
        print("\n按任意键退出...")
        input()
