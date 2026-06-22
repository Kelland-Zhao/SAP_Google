import warnings
warnings.filterwarnings('ignore', category=FutureWarning)

import subprocess
import time
import win32com.client
import os
import datetime
import calendar
import shutil
import glob
import openpyxl
import gspread
import sys
import requests
from google.auth.transport.requests import AuthorizedSession
from google.oauth2 import service_account

OUTPUT_DIR = r"O:\My Drive\071 - SAP 数据\Maintenance_Plan_Adherence"
SAP_DEFAULT_DOWNLOAD_DIR = r"O:\My Drive\071 - SAP 数据\Maintenance_Plan_Adherence"

GOOGLE_SHEET_ID = '1dWWNiIlRqnXDeDiY4MS_cPSfseDb0g1NHK1dC1HPSnY'
WORKSHEET_NAME = 'MasterData'

def get_resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

SERVICE_ACCOUNT_FILE = get_resource_path('../pyreadsp-b5b9c1909de6.json')

def sap_auto_logo():
    subprocess.check_call(['C:\\Program Files (x86)\\SAP\\FrontEnd\\SAPgui\\sapshcut.exe', '-system=CAP', '-client=321',
                           '-user=USERNAME', '-pw=PASSWORD', '-language=ZH'])
    time.sleep(15)
    print("sap open successfully")


def close_SAP():
    try:
        os.system('taskkill /im saplogon.exe /t /f')
        print("SAP GUI 进程已成功关闭。")
    except Exception as e:
        print(f"关闭 SAP 进程失败: {e}")


def process_maintenance_data(excel_file_path, year_month):
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active
    
    system_status_col = 11
    header_row = 1
    
    header_value = ws.cell(header_row, system_status_col).value
    if header_value != "系统字段":
        print(f"警告：K 列的列名是 '{header_value}'，预期是 '系统字段'")
    
    total_planned = 0
    executed_count = 0
    
    for row in range(2, ws.max_row + 1):
        status = ws.cell(row, system_status_col).value
        if status and str(status).strip():
            total_planned += 1
            if "TECO" in str(status).upper():
                executed_count += 1
    
    wb.close()
    
    adherence = round((executed_count / total_planned), 4) if total_planned > 0 else 0
    
    return {
        'month': year_month,
        'executed': executed_count,
        'planned': total_planned,
        'adherence_pct': adherence
    }


def upload_to_google_sheets(data, sheet_id, worksheet_name, auth_file):
    try:
        print("正在连接 Google Sheets...")
        
        # 创建禁用SSL验证的session
        session = requests.Session()
        session.verify = False
        
        # 禁用SSL警告
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        
        # 使用自定义session创建gspread客户端
        credentials = service_account.Credentials.from_service_account_file(
            auth_file,
            scopes=['https://www.googleapis.com/auth/spreadsheets']
        )
        authed_session = AuthorizedSession(credentials)
        authed_session.verify = False
        
        gc = gspread.Client(auth=credentials, session=authed_session)
        sh = gc.open_by_key(sheet_id)
        worksheet = sh.worksheet(worksheet_name)
        
        all_values = worksheet.get_all_values()
        month_col_index = 0
        
        target_row = None
        for idx, row in enumerate(all_values):
            if len(row) > 0 and row[month_col_index] == data['month']:
                target_row = idx + 1
                break
        
        update_data = [
            data['month'],
            data['executed'],
            data['planned'],
            data['adherence_pct']
        ]
        
        if target_row:
            worksheet.update(f'A{target_row}:D{target_row}', [update_data])
            print(f"✅ 已更新月份 {data['month']} 的数据（第 {target_row} 行）")
        else:
            worksheet.append_row(update_data, value_input_option='USER_ENTERED')
            print(f"✅ 已添加月份 {data['month']} 的新行")
        
        return True
        
    except gspread.exceptions.WorksheetNotFound:
        print(f"❌ Google Sheets 错误: 找不到工作表 '{worksheet_name}'。")
        return False
    except gspread.exceptions.SpreadsheetNotFound:
        print(f"❌ Google Sheets 错误: 找不到工作簿。请确认已将服务账户邮箱添加为该表格的 '编辑者'。")
        return False
    except Exception as e:
        print(f"❌ Google Sheets 上传失败。错误: {e}")
        import traceback
        traceback.print_exc()
        return False


def get_maintenance_plan_iw39():
    today = datetime.date.today()
    first_day = today.replace(day=1)
    last_day = today.replace(day=calendar.monthrange(today.year, today.month)[1])
    
    first_day_str = first_day.strftime("%m/%d/%Y")
    last_day_str = last_day.strftime("%m/%d/%Y")
    
    year_month = today.strftime("%Y%m")
    output_filename = f"{year_month}_Maintenance_Plan_Adherence.xlsx"
    output_file_path = os.path.join(OUTPUT_DIR, output_filename)
    
    if os.path.exists(output_file_path):
        try:
            os.remove(output_file_path)
            print(f"已删除旧文件: {output_file_path}")
        except Exception as e:
            print(f"警告: 无法删除旧文件 {output_file_path}: {e}")
    
    try:
        SapGuiAuto = win32com.client.GetObject("SAPGUI")
        application = SapGuiAuto.GetScriptingEngine
        connection = application.Children(0)
        session = connection.Children(0)
    except Exception as e:
        print(f"错误: 无法连接到 SAP GUI Scripting Engine 或找不到活动会话。请确保 SAP GUI 已登录。{e}")
        return
    
    print("成功连接到 SAP 会话。正在执行 IW39...")
    
    try:
        session.findById("wnd[0]").maximize()
        session.findById("wnd[0]/tbar[0]/okcd").text = "IW39"
        session.findById("wnd[0]").sendVKey(0)
        
        session.findById("wnd[0]/usr/chkDY_MAB").selected = True
        session.findById("wnd[0]/usr/chkDY_IAR").selected = True
        session.findById("wnd[0]/usr/chkDY_IAR").setFocus()
        session.findById("wnd[0]").sendVKey(2)
        
        session.findById("wnd[0]/usr/chkDY_OFN").selected = True
        session.findById("wnd[0]/usr/chkDY_OFN").setFocus()
        session.findById("wnd[0]").sendVKey(2)
        
        session.findById("wnd[0]/usr/ctxtAUART-LOW").setFocus()
        session.findById("wnd[0]/usr/ctxtAUART-LOW").caretPosition = 0
        session.findById("wnd[0]/usr/btn%_AUART_%_APP_%-VALU_PUSH").press()
        
        session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,0]").text = "PM02"
        session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,1]").text = "PM03"
        session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,2]").text = "PM10"
        session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,3]").text = "ZPM3"
        session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,4]").text = "ZPM4"
        session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,5]").text = "ZPM8"
        session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,5]").setFocus()
        session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,5]").caretPosition = 4
        print("工单类型筛选: PM02, PM03, PM10, ZPM3, ZPM4, ZPM8")
        session.findById("wnd[1]/tbar[0]/btn[8]").press()
        
        session.findById("wnd[0]/usr/ctxtSTRNO-LOW").text = "CN15*"
        session.findById("wnd[0]/usr/ctxtIWERK-LOW").text = "CN15"
        session.findById("wnd[0]/usr/ctxtGLTRS-LOW").setFocus()
        session.findById("wnd[0]/usr/ctxtGLTRS-LOW").caretPosition = 0
        session.findById("wnd[0]").sendVKey(4)
        
        session.findById("wnd[1]/usr/cntlCONTAINER/shellcont/shell").focusDate = first_day.strftime("%Y%m%d")
        session.findById("wnd[1]/usr/cntlCONTAINER/shellcont/shell").firstVisibleDate = first_day.strftime("%Y%m%d")
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        
        session.findById("wnd[0]/usr/ctxtGLTRS-LOW").text = first_day_str
        session.findById("wnd[0]/usr/ctxtGLTRS-HIGH").text = last_day_str
        session.findById("wnd[0]/usr/ctxtGLTRS-HIGH").setFocus()
        session.findById("wnd[0]/usr/ctxtGLTRS-HIGH").caretPosition = 10
        session.findById("wnd[0]").sendVKey(2)
        
        session.findById("wnd[1]/tbar[0]/btn[12]").press()
        
        session.findById("wnd[0]/usr/ctxtGLTRS-LOW").text = first_day_str
        session.findById("wnd[0]/usr/ctxtGLTRS-HIGH").text = last_day_str
        session.findById("wnd[0]/usr/ctxtSWERK-LOW").text = "CN15"
        session.findById("wnd[0]/usr/ctxtSWERK-LOW").setFocus()
        session.findById("wnd[0]/usr/ctxtSWERK-LOW").caretPosition = 4
        
        session.findById("wnd[0]/usr/ctxtVARIANT").text = "/KEL"
        session.findById("wnd[0]/usr/ctxtVARIANT").setFocus()
        session.findById("wnd[0]/usr/ctxtVARIANT").caretPosition = 4
        
        session.findById("wnd[0]/tbar[1]/btn[8]").press()
        time.sleep(2)
        
        session.findById("wnd[0]/mbar/menu[0]/menu[6]").select()
        time.sleep(1)
        
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        
        session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[0,0]").select()
        session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[0,0]").setFocus()
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        
        print("等待 Excel 打开...")
        time.sleep(5)
        
        print("正在连接到 Excel 应用程序...")
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            
            if excel.Workbooks.Count > 0:
                wb = excel.Workbooks(1)
                print(f"找到打开的工作簿: {wb.Name}")
                
                if not os.path.exists(OUTPUT_DIR):
                    os.makedirs(OUTPUT_DIR)
                    print(f"创建目录: {OUTPUT_DIR}")
                
                print(f"正在另存为到: {output_file_path}")
                wb.SaveAs(output_file_path)
                
                wb.Close(SaveChanges=False)
                excel.Quit()
                
                print(f"✅ IW39 维护计划数据导出完成")
                print(f"文件保存路径: {output_file_path}")
                print(f"日期范围: {first_day_str} - {last_day_str}")
                
                print("\n" + "="*50)
                print("开始处理数据并上传到 Google Sheets...")
                print("="*50)
                
                data = process_maintenance_data(output_file_path, year_month)
                print(f"\n统计结果：")
                print(f"  - 月份: {data['month']}")
                print(f"  - 已执行订单数: {data['executed']}")
                print(f"  - 计划订单总数: {data['planned']}")
                print(f"  - 维护计划遵守率: {data['adherence_pct']}")
                
                upload_to_google_sheets(
                    data=data,
                    sheet_id=GOOGLE_SHEET_ID,
                    worksheet_name=WORKSHEET_NAME,
                    auth_file=SERVICE_ACCOUNT_FILE
                )
                
            else:
                print("⚠️ 警告: 没有找到打开的 Excel 工作簿")
        except Exception as e:
            print(f"❌ 处理 Excel 时发生错误: {e}")
            import traceback
            traceback.print_exc()
        
        time.sleep(2)
        
    except Exception as e:
        print(f"❌ 执行 SAP 操作时发生错误: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    try:
        print("正在清理可能存在的 SAP 进程...")
        close_SAP()
        time.sleep(2)
        
        print("正在启动 SAP GUI...")
        sap_auto_logo()
        
        print("等待 SAP GUI 完全加载...")
        time.sleep(5)
        
        print("开始执行 SAP 操作...")
        get_maintenance_plan_iw39()
        
        print("\nSAP 操作完成，正在关闭 SAP...")
        time.sleep(2)
        close_SAP()
        
        print("\n" + "="*50)
        print("✅ 所有操作已完成！")
        print("="*50)
        
    except Exception as e:
        print(f"❌ 程序执行过程中发生错误: {e}")
        import traceback
        traceback.print_exc()
        close_SAP()
    finally:
        print("\n按任意键退出...")
        input()
