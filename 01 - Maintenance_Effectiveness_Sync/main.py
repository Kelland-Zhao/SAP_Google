import subprocess
import time
import win32com.client
import os
import datetime
import calendar
import openpyxl
import gspread
import sys
import requests
import warnings

OUTPUT_DIR = r"O:\My Drive\071 - SAP 数据\Maintenance_Effectiveness"

GOOGLE_SHEET_ID = '1dWWNiIlRqnXDeDiY4MS_cPSfseDb0g1NHK1dC1HPSnY'
WORKSHEET_NAME = 'MasterData'

def get_resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

SERVICE_ACCOUNT_FILE = get_resource_path('pyreadsp-b5b9c1909de6.json')

def sap_auto_logo():
    subprocess.check_call(['C:\\Program Files (x86)\\SAP\\FrontEnd\\SAPgui\\sapshcut.exe', '-system=CAP', '-client=321',
                           '-user=USERNAME', '-pw=PASSWORD', '-language=ZH'])
    time.sleep(15)
    print("sap open successfully")


def close_SAP():
    try:
        os.system('taskkill /im saplogon.exe /t /f')
        print("SAP GUI 进程已成功关闭。")
    except Exception as e:
        print(f"关闭 SAP 进程失败: {e}")


def process_effectiveness_data(excel_file_path, year_month):
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active
    
    header_row = 1
    system_status_col = None
    order_type_col = None
    actual_work_col = None
    
    for col in range(1, ws.max_column + 1):
        header_value = ws.cell(header_row, col).value
        if header_value == "系统状态":
            system_status_col = col
        elif header_value == "订单类型":
            order_type_col = col
        elif header_value == "实际的工作":
            actual_work_col = col
    
    if not all([system_status_col, order_type_col, actual_work_col]):
        print(f"警告：未找到所需列。系统状态: {system_status_col}, 订单类型: {order_type_col}, 实际的工作: {actual_work_col}")
        wb.close()
        return None
    
    total_hours = 0.0
    planned_hours = 0.0
    
    for row in range(2, ws.max_row + 1):
        status = ws.cell(row, system_status_col).value
        order_type = ws.cell(row, order_type_col).value
        actual_work = ws.cell(row, actual_work_col).value
        
        if status and "CNF" in str(status).upper():
            try:
                work_value = float(actual_work) if actual_work else 0.0
                total_hours += work_value
                
                if order_type and str(order_type).strip() != "PM01":
                    planned_hours += work_value
            except (ValueError, TypeError):
                continue
    
    wb.close()
    
    total_hours = round(total_hours, 1)
    planned_hours = round(planned_hours, 1)
    effectiveness = round((planned_hours / total_hours), 4) if total_hours > 0 else 0.0
    
    return {
        'month': year_month,
        'planned_hours': planned_hours,
        'total_hours': total_hours,
        'effectiveness_pct': effectiveness
    }


def upload_to_google_sheets(data, sheet_id, worksheet_name, auth_file):
    try:
        print("正在连接 Google Sheets...")
        warnings.filterwarnings('ignore', message='Unverified HTTPS request')
        urllib3 = requests.packages.urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        
        session = requests.Session()
        session.verify = False
        
        from google.auth.transport.requests import AuthorizedSession
        from google.oauth2.service_account import Credentials
        
        scopes = [
            'https://www.googleapis.com/auth/spreadsheets',
            'https://www.googleapis.com/auth/drive'
        ]
        credentials = Credentials.from_service_account_file(auth_file, scopes=scopes)
        
        authed_session = AuthorizedSession(credentials)
        authed_session.verify = False
        
        gc = gspread.Client(auth=credentials, session=authed_session)
        sh = gc.open_by_key(sheet_id)
        worksheet = sh.worksheet(worksheet_name)
        
        all_values = worksheet.get_all_values()
        month_col_index = 0
        
        target_row = None
        for idx, row in enumerate(all_values):
            if len(row) > 0 and row[month_col_index] == data['month']:
                target_row = idx + 1
                break
        
        if target_row:
            worksheet.update_cell(target_row, 5, data['planned_hours'])
            worksheet.update_cell(target_row, 6, data['total_hours'])
            worksheet.update_cell(target_row, 7, data['effectiveness_pct'])
            print(f"✅ 已更新月份 {data['month']} 的数据（第 {target_row} 行，E-G 列）")
        else:
            update_data = [data['month'], '', '', '', data['planned_hours'], data['total_hours'], data['effectiveness_pct']]
            worksheet.append_row(update_data, value_input_option='USER_ENTERED')
            print(f"✅ 已添加月份 {data['month']} 的新行（E-G 列）")
        
        return True
        
    except gspread.exceptions.WorksheetNotFound:
        print(f"❌ Google Sheets 错误: 找不到工作表 '{worksheet_name}'。")
        return False
    except gspread.exceptions.SpreadsheetNotFound:
        print(f"❌ Google Sheets 错误: 找不到工作簿。请确认已将服务账户邮箱添加为该表格的 '编辑者'。")
        return False
    except Exception as e:
        print(f"❌ Google Sheets 上传失败。错误: {e}")
        import traceback
        traceback.print_exc()
        return False


def get_maintenance_effectiveness_iw47():
    today = datetime.date.today()
    first_day = today.replace(day=1)
    last_day = today.replace(day=calendar.monthrange(today.year, today.month)[1])
    
    first_day_str = first_day.strftime("%m/%d/%Y")
    last_day_str = last_day.strftime("%m/%d/%Y")
    
    year_month = today.strftime("%Y%m")
    output_filename = f"{year_month}_Maintenance_Effectiveness.xlsx"
    output_file_path = os.path.join(OUTPUT_DIR, output_filename)
    
    if os.path.exists(output_file_path):
        try:
            os.remove(output_file_path)
            print(f"已删除旧文件: {output_file_path}")
        except Exception as e:
            print(f"警告: 无法删除旧文件 {output_file_path}: {e}")
    
    try:
        SapGuiAuto = win32com.client.GetObject("SAPGUI")
        application = SapGuiAuto.GetScriptingEngine
        connection = application.Children(0)
        session = connection.Children(0)
    except Exception as e:
        print(f"错误: 无法连接到 SAP GUI Scripting Engine 或找不到活动会话。请确保 SAP GUI 已登录。{e}")
        return
    
    print("成功连接到 SAP 会话。正在执行 IW47...")
    
    try:
        session.findById("wnd[0]").maximize()
        session.findById("wnd[0]/tbar[0]/okcd").text = "IW47"
        session.findById("wnd[0]").sendVKey(0)
        time.sleep(1)
        
        session.findById("wnd[0]/usr/chkDY_IAR").selected = True
        session.findById("wnd[0]/usr/chkDY_ABG").selected = True
        session.findById("wnd[0]/usr/chkDY_ABG").setFocus()
        
        session.findById("wnd[0]/usr/btn%_AUART_O_%_APP_%-VALU_PUSH").press()
        time.sleep(1)
        
        print("正在设置工单类型筛选条件: PM01, PM02, PM03, ZPM3, ZPM4, PM10, ZPM8")
        session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,0]").text = "PM01"
        session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,1]").text = "PM02"
        session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,2]").text = "PM03"
        session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,3]").text = "ZPM3"
        session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,4]").text = "ZPM4"
        session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,5]").text = "PM10"
        session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,6]").text = "ZPM8"
        session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,6]").setFocus()
        session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,6]").caretPosition = 4
        session.findById("wnd[1]/tbar[0]/btn[8]").press()
        time.sleep(1)
        
        session.findById("wnd[0]/usr/ctxtERSDA_C-LOW").text = first_day_str
        session.findById("wnd[0]/usr/ctxtERSDA_C-HIGH").setFocus()
        session.findById("wnd[0]/usr/ctxtERSDA_C-HIGH").caretPosition = 10
        session.findById("wnd[0]").sendVKey(2)
        time.sleep(1)
        
        session.findById("wnd[1]/tbar[0]/btn[12]").press()
        time.sleep(1)
        
        session.findById("wnd[0]/usr/ctxtERSDA_C-LOW").text = first_day_str
        session.findById("wnd[0]/usr/ctxtERSDA_C-HIGH").text = last_day_str
        session.findById("wnd[0]/usr/ctxtWERKS_C-LOW").text = "CN15"
        session.findById("wnd[0]/usr/ctxtVARIANT").text = "/BADDI 15"
        session.findById("wnd[0]/usr/ctxtVARIANT").setFocus()
        session.findById("wnd[0]/usr/ctxtVARIANT").caretPosition = 9
        
        session.findById("wnd[0]/tbar[1]/btn[8]").press()
        time.sleep(3)
        
        session.findById("wnd[0]/mbar/menu[0]/menu[6]").select()
        time.sleep(1)
        
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        
        session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[0,0]").select()
        session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[0,0]").setFocus()
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        
        print("等待 Excel 打开...")
        time.sleep(5)
        
        print("正在连接到 Excel 应用程序...")
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            
            if excel.Workbooks.Count > 0:
                wb = excel.Workbooks(1)
                print(f"找到打开的工作簿: {wb.Name}")
                
                if not os.path.exists(OUTPUT_DIR):
                    os.makedirs(OUTPUT_DIR)
                    print(f"创建目录: {OUTPUT_DIR}")
                
                print(f"正在另存为到: {output_file_path}")
                wb.SaveAs(output_file_path)
                
                wb.Close(SaveChanges=False)
                excel.Quit()
                
                print(f"✅ IW47 维护有效性数据导出完成")
                print(f"文件保存路径: {output_file_path}")
                print(f"日期范围: {first_day_str} - {last_day_str}")
                
                print("\n" + "="*50)
                print("开始处理数据并上传到 Google Sheets...")
                print("="*50)
                
                data = process_effectiveness_data(output_file_path, year_month)
                if data:
                    print(f"\n统计结果：")
                    print(f"  - 月份: {data['month']}")
                    print(f"  - 已完成计划工单工时: {data['planned_hours']}")
                    print(f"  - 所有已完成工单工时: {data['total_hours']}")
                    print(f"  - 维护有效性: {data['effectiveness_pct']}")
                    
                    upload_to_google_sheets(
                        data=data,
                        sheet_id=GOOGLE_SHEET_ID,
                        worksheet_name=WORKSHEET_NAME,
                        auth_file=SERVICE_ACCOUNT_FILE
                    )
                else:
                    print("❌ 数据处理失败")
                
            else:
                print("⚠️ 警告: 没有找到打开的 Excel 工作簿")
        except Exception as e:
            print(f"❌ 处理 Excel 时发生错误: {e}")
            import traceback
            traceback.print_exc()
        
        time.sleep(2)
        
    except Exception as e:
        print(f"❌ 执行 SAP 操作时发生错误: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    try:
        print("正在清理可能存在的 SAP 进程...")
        close_SAP()
        time.sleep(2)
        
        print("正在启动 SAP GUI...")
        sap_auto_logo()
        
        print("等待 SAP GUI 完全加载...")
        time.sleep(5)
        
        print("开始执行 IW47 维护有效性数据提取...")
        get_maintenance_effectiveness_iw47()
        
        print("\nSAP 操作完成，正在关闭 SAP...")
        close_SAP()
        
        print("\n" + "="*50)
        print("✅ 所有操作已完成！")
        print("="*50)
        
    except Exception as e:
        print(f"❌ 程序执行过程中发生错误: {e}")
        import traceback
        traceback.print_exc()
        close_SAP()
    finally:
        print("\n按任意键退出...")
        input()
